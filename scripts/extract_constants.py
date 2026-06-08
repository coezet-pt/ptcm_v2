#!/usr/bin/env python3
"""
Extract / audit constants from CoEZET_PTCM_v3.xlsx.

Produces:
  - scripts/extracted_audit.json  (structured dump of everything we read)
  - console output: bucket->segment map, sum-check, BAU reference table
"""
import json
import sys
from pathlib import Path
import openpyxl

WB_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/v3.xlsx")
OUT = Path(__file__).parent / "extracted_audit.json"

wb = openpyxl.load_workbook(WB_PATH, data_only=True)


# --------------------------------------------------------------------------
# 1) Bucket -> segment (vehicle size) and bucket -> application (use case)
# --------------------------------------------------------------------------
ws = wb["Buckets"]
buckets = []
for r in range(3, 17):  # B1..B14
    bid = ws.cell(r, 2).value
    usecase = ws.cell(r, 3).value
    model = ws.cell(r, 4).value
    v2025 = ws.cell(r, 5).value
    share2045 = ws.cell(r, 19).value  # 'Share of TIV' for 2045
    if not bid:
        continue
    buckets.append({
        "id": bid,
        "application": (usecase or "").strip(),
        "size": (model or "").strip(),
        "vol_2025": v2025,
        "tiv_share_2045": share2045,
    })

# Vehicle-size families used as the workbook's "Segment" axis
def size_family(s: str) -> str:
    s = s.lower()
    if "tip" in s: return "Tipper (Rigid)"
    if "tractor" in s or "tt" in s: return "Tractor (T-T)"
    return "Rigid"

for b in buckets:
    b["segment"] = size_family(b["size"])

SEGMENTS = sorted({b["segment"] for b in buckets})
APPLICATIONS = sorted({b["application"] for b in buckets})


# --------------------------------------------------------------------------
# 2) Steady-state share sum-check across the 5 Estimation sheets
#    Each bucket block is 12 rows; "Power Train Share (Final)" at offset +10
#    Shares in columns R..W = 18..23 (Diesel, CNG, LNG, BET, H2-ICE, H2-FCET)
# --------------------------------------------------------------------------
EST_SHEETS = {
    2035: "Estimation2035",
    2040: "Estimation2040",
    2045: "Estimation SS2045",
    2050: "Estimation2050",
    2055: "Estimation 100% ZET 2055",
}
PT_ORDER = ["Diesel", "CNG", "LNG", "BET", "H2-ICE", "H2-FCET"]

steady_state = {}      # year -> bucket -> {pt: share}
sum_violations = []

for year, sn in EST_SHEETS.items():
    w = wb[sn]
    steady_state[year] = {}
    for i in range(14):  # 14 buckets
        # Final-share row: row 15, 27, 39, ...
        row = 15 + i * 12
        bid = f"B{i+1}"
        shares = {}
        for j, pt in enumerate(PT_ORDER):
            v = w.cell(row, 18 + j).value
            shares[pt] = float(v) if isinstance(v, (int, float)) else 0.0
        s = sum(shares.values())
        steady_state[year][bid] = shares
        if abs(s - 1.0) > 0.005 and s > 0:  # tolerance 0.5%
            sum_violations.append((year, bid, round(s, 4), shares))


# --------------------------------------------------------------------------
# 3) Output Summary BAU sales reference (rows 29-59 = years 2025-2055)
#    cols: B=Diesel sale, D=BET, F=H2-ICE, H=FCET, J=CNG, L=LNG, N=Total
# --------------------------------------------------------------------------
ws = wb["Output Summary"]
# Positional map — workbook headers are buggy; map by column index.
SALE_COL  = {"Diesel": 2, "BET": 4, "H2-ICE": 6, "H2-FCET": 8, "CNG": 10, "LNG": 12}
STOCK_COL = {"Diesel": 3, "BET": 5, "H2-ICE": 7, "H2-FCET": 9, "CNG": 11, "LNG": 13}
TOTAL_SALE_COL = 14
bau_ref = {}
for r in range(29, 60):
    year = ws.cell(r, 1).value
    row = {pt: (ws.cell(r, c).value or 0) for pt, c in SALE_COL.items()}
    row["Total"] = ws.cell(r, TOTAL_SALE_COL).value or 0
    for pt, c in STOCK_COL.items():
        row[f"{pt}_stock"] = ws.cell(r, c).value or 0
    bau_ref[year] = row

output_summary_headers = {
    "row1": [ws.cell(1, c).value for c in range(1, 15)],
    "row2": [ws.cell(2, c).value for c in range(1, 15)],
}


# --------------------------------------------------------------------------
# 4) TIV projections per year (Buckets sheet) for each Estimation year
#    From Buckets row 18 (Total) - column H=2035, M=2040, R=2045, W=2050, AB=2055
# --------------------------------------------------------------------------
ws = wb["Buckets"]
tiv_anchors = {
    2035: ws.cell(2, 8).value,    # header lives in row 2 (Potential TIV 2035 total — actually below)
}
# Easier: read from Estimation sheets (row 2, col 5 has TIV estimate)
tiv_estimates = {}
for year, sn in EST_SHEETS.items():
    tiv_estimates[year] = wb[sn].cell(2, 5).value


# --------------------------------------------------------------------------
# Dump audit JSON
# --------------------------------------------------------------------------
audit = {
    "buckets": buckets,
    "segments": SEGMENTS,
    "applications": APPLICATIONS,
    "steady_state_shares": steady_state,
    "sum_violations": sum_violations,
    "bau_reference": bau_ref,
    "tiv_estimates": tiv_estimates,
    "output_summary_headers": output_summary_headers,
}
OUT.write_text(json.dumps(audit, indent=2, default=str))


# --------------------------------------------------------------------------
# CONSOLE REPORT
# --------------------------------------------------------------------------
print("=" * 72)
print("(a) BUCKET → SEGMENT / APPLICATION MAP")
print("=" * 72)
print(f"{'Bucket':<6} {'Size':<14} {'Segment':<18} {'Application':<32} {'Share2045':>9}")
for b in buckets:
    print(f"{b['id']:<6} {b['size']:<14} {b['segment']:<18} {b['application']:<32} {b['tiv_share_2045']:>9.4f}")
print(f"\nDistinct segments ({len(SEGMENTS)}): {SEGMENTS}")
print(f"Distinct applications ({len(APPLICATIONS)}): {APPLICATIONS}")

print("\n" + "=" * 72)
print("(b) STEADY-STATE SHARE SUM CHECK (5 sheets × 14 buckets × 6 PTs)")
print("=" * 72)
if not sum_violations:
    print("✅ all steady-state shares sum to 1.0 ±0.5%")
else:
    print(f"⚠️  {len(sum_violations)} bucket/year combos NOT summing to 1.0:")
    for year, bid, s, shares in sum_violations:
        print(f"   {year} {bid}  sum={s}  shares={shares}")
print(f"\nTIV anchors from Estimation sheets: {tiv_estimates}")

print("\n" + "=" * 72)
print("(c) BAU REFERENCE — Output Summary rows 29–59")
print("=" * 72)
print(f"{'Year':<6} {'Diesel':>10} {'CNG':>9} {'LNG':>7} {'BET':>10} {'H2ICE':>9} {'FCET':>9} {'Total':>10}")
for year, row in sorted(bau_ref.items()):
    print(f"{year:<6} {row['Diesel']:>10,.0f} {row['CNG']:>9,.0f} {row['LNG']:>7,.0f} "
          f"{row['BET']:>10,.0f} {row['H2-ICE']:>9,.0f} {row['H2-FCET']:>9,.0f} {row['Total']:>10,.0f}")

print("\n✅ Wrote audit to", OUT)
